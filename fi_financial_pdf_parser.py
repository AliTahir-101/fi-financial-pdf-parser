import tabula
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

def extract_table_data(pdf_path: str, template_path: str):
    """
    Simple function to extract data from a PDF using a template and save to Excel.
    
    Args:
        pdf_path: Path to the PDF file
        template_path: Path to the tabula template JSON file
    """
    try:
        print(f"Reading PDF: {pdf_path}")
        print(f"Using template: {template_path}")
        
        # Extract tables from PDF using the template
        tables = tabula.read_pdf_with_template(
            pdf_path,
            template_path,
            stream=True,
        )
        
        print(f"Found {len(tables)} table(s)")
        
        if not tables:
            print("No tables found in the PDF")
            return
        
        # Clean the tables and combine them
        cleaned_tables = []
        # Remove empty rows and columns
        cleaned_table = tables[0].dropna(how='all').dropna(axis=1, how='all')

        # Clean string data
        for col in cleaned_table.columns:
            if cleaned_table[col].dtype == 'object':
                cleaned_table[col] = cleaned_table[col].astype(str).str.strip()
                cleaned_table[col] = cleaned_table[col].replace('nan', '')
        
        if not cleaned_table.empty:
            print(f"Table shape: {cleaned_table.shape}")
            return cleaned_table

        if not cleaned_tables:
            print("No valid data found in tables")
            return
        
    except Exception as e:
        print(f"Error processing PDF: {e}")


def extract_table_name(pdf_path: str, template_path: str) -> str:
    """
    Extracts the table name from the PDF using the specified template.
    Args:
        pdf_path: Path to the PDF file
        template_path: Path to the tabula template JSON file
    Returns:
        The extracted table name as a string.
    """
    try:
        print(f"Extracting table name from PDF: {pdf_path}")
        print(f"Using template: {template_path}")
        
        # Extract tables from PDF using the template
        tables = tabula.read_pdf_with_template(
            pdf_path,
            template_path,
            stream=True,
        )
        print(f"Found {len(tables)} table(s) for table name extraction")
        if not tables or len(tables) == 0:
            print("No tables found for table name extraction")
            return ""
        
        # Assuming the first table contains the table name
        table_name_df = tables[0]
        return table_name_df.columns[0].strip() if table_name_df.columns.size > 0 else ""
    
    except Exception as e:
        print(f"Error extracting table name: {e}")
        return ""


def save_tables_to_excel(tables_data: list, output_path: str):
    """
    Save multiple tables to Excel with formatting:
    - Table names highlighted in green
    - Column headers highlighted in yellow
    - 5 rows gap between each table
    
    Args:
        tables_data: List of tuples (table_name, table_dataframe)
        output_path: Path where the Excel file will be saved
    """
    try:
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Define styles
        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        
        current_row = 1
        
        for table_idx, (table_name, table_data) in enumerate(tables_data):
            print(f"Adding table {table_idx + 1}: {table_name}")
            
            # Add table name with green highlighting
            if table_name:
                ws.cell(row=current_row, column=1, value=table_name)
                ws.cell(row=current_row, column=1).fill = green_fill
                ws.cell(row=current_row, column=1).font = bold_font
                current_row += 2  # Leave a blank row after table name
            
            # Add column headers with yellow highlighting
            for col_idx, column_name in enumerate(table_data.columns, 1):
                if "Unnamed" in str(column_name):
                    column_name = ""
                cell = ws.cell(row=current_row, column=col_idx, value=str(column_name))
                cell.fill = yellow_fill
                cell.font = bold_font
            
            current_row += 1
            
            # Add data rows
            for row_data in dataframe_to_rows(table_data, index=False, header=False):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=current_row, column=col_idx, value=value)
                current_row += 1
            
            # Add 5 rows gap between tables (except after the last table)
            if table_idx < len(tables_data) - 1:
                current_row += 5
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(output_path)
        print(f"Successfully saved {len(tables_data)} tables to: {output_path}")
        
    except Exception as e:
        print(f"Error saving to Excel: {e}")

def main():
    # Configuration - change these paths as needed
    pdf_file = "./input_pdf_files/new_format/test.pdf"
    template_dir = "./templates/new_format"
    output_file = "./output.xlsx"
    num_templates = 69  # Define how many templates to process

    # Check if PDF file exists
    if not Path(pdf_file).exists():
        print(f"PDF file not found: {pdf_file}")
        return

    # List to store all extracted tables
    all_tables = []
    
    print(f"Processing {num_templates} templates...")
    
    for template_number in range(1, num_templates + 1):
        print(f"\n--- Processing Template {template_number} ---")
        
        # Extract Table Name
        table_name_template_file = f"{template_dir}/{template_number}_table_name.tabula-template.json"
        table_name = ""
        if Path(table_name_template_file).exists():
            table_name = extract_table_name(pdf_file, table_name_template_file)
            print(f"Table Name: {table_name}")
        else:
            print(f"Table name template not found: {table_name_template_file}")
            table_name = f"Table {template_number}"  # Default name with number
        
        # Extract data
        table_data_template_file = f"{template_dir}/{template_number}_table_data.tabula-template.json"
        if not Path(table_data_template_file).exists():
            print(f"Table data template not found: {table_data_template_file}")
            # Try fallback to numbered template without separate name/data files
            fallback_template = f"{template_dir}/{template_number}.tabula-template.json"
            if Path(fallback_template).exists():
                table_data_template_file = fallback_template
                print(f"Using fallback template: {fallback_template}")
            else:
                print(f"Skipping template {template_number} - no template file found")
                continue
        
        table_data = extract_table_data(pdf_file, table_data_template_file)
        
        if table_data is not None and not table_data.empty:
            print(f"Extracted table with shape: {table_data.shape}")
            print("First few rows:")
            print(table_data.head(3))
            
            # Add to our collection of tables
            all_tables.append((table_name, table_data))
        else:
            print(f"No valid table data extracted for template {template_number}")
    
    # Save all tables to Excel with formatting and gaps
    if all_tables:
        print(f"\nSaving {len(all_tables)} tables to Excel...")
        save_tables_to_excel(all_tables, output_file)
    else:
        print("No tables were successfully extracted")

if __name__ == "__main__":
    main()
